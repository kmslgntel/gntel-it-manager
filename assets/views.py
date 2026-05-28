import io
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Count
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import IpAddress, IpGroup, Person, Phone, Company
from .forms import IpEditForm
from core.utils import log_action, ip_to_int


# ──────────────────────────────────────────────
# IP 관리
# ──────────────────────────────────────────────

@login_required
def ip_list(request):
    qs = IpAddress.objects.select_related(
        'group', 'person', 'person__department', 'person__company'
    ).order_by('ip_int')

    q = request.GET.get('q', '').strip()
    group_id = request.GET.get('group', '')
    status = request.GET.get('status', '')
    company_id = request.GET.get('company', '')

    if q:
        qs = qs.filter(
            Q(ip__icontains=q) |
            Q(person__name__icontains=q) |
            Q(note__icontains=q)
        )
    if group_id:
        qs = qs.filter(group_id=group_id)
    if status == 'used':
        qs = qs.filter(person__isnull=False)
    elif status == 'free':
        qs = qs.filter(person__isnull=True)
    if company_id:
        qs = qs.filter(person__company_id=company_id)

    ip_total = IpAddress.objects.count()
    ip_used = IpAddress.objects.filter(person__isnull=False).count()
    ip_free = ip_total - ip_used

    # 회사별 IP 사용 현황 (person이 있는 것만)
    company_stats = (
        IpAddress.objects
        .filter(person__company__isnull=False)
        .values('person__company__id', 'person__company__name')
        .annotate(used=Count('pk'))
        .order_by('-used')
    )

    paginator = Paginator(qs, 100)
    page_obj = paginator.get_page(request.GET.get('page'))

    # 그룹을 parent_name 기준으로 묶어서 전달
    groups_raw = IpGroup.objects.filter(is_phone_group=False).order_by('parent_name', 'name')
    groups_by_parent = {}
    for g in groups_raw:
        parent = g.parent_name or '기타'
        groups_by_parent.setdefault(parent, []).append(g)

    companies = Company.objects.order_by('name')

    return render(request, 'assets/ip_list.html', {
        'page_obj': page_obj,
        'q': q,
        'group_id': group_id,
        'status': status,
        'company_id': company_id,
        'groups_by_parent': groups_by_parent,
        'companies': companies,
        'company_stats': company_stats,
        'ip_total': ip_total,
        'ip_used': ip_used,
        'ip_free': ip_free,
    })


@login_required
def ip_detail(request, pk):
    obj = get_object_or_404(
        IpAddress.objects.select_related('group', 'person', 'person__department', 'person__company'),
        pk=pk
    )
    from core.models import AuditLog
    logs = AuditLog.objects.filter(target_type='IpAddress', target_id=str(pk)).order_by('-acted_at')[:20]
    return render(request, 'assets/ip_detail.html', {'obj': obj, 'logs': logs})


@login_required
def ip_update(request, pk):
    obj = get_object_or_404(IpAddress, pk=pk)
    if request.method == 'POST':
        form = IpEditForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            log_action(request, 'update', 'IpAddress', obj.pk,
                       f'IP 수정: {obj.ip}')
            messages.success(request, f'IP {obj.ip}가 수정되었습니다.')
            return redirect('assets:ip_detail', pk=obj.pk)
    else:
        form = IpEditForm(instance=obj)
    return render(request, 'assets/ip_form.html', {'form': form, 'obj': obj})


@login_required
def ip_delete(request, pk):
    obj = get_object_or_404(IpAddress, pk=pk)
    if request.method == 'POST':
        ip_str = obj.ip
        obj.delete()
        log_action(request, 'delete', 'IpAddress', pk, f'IP 삭제: {ip_str}')
        messages.success(request, f'IP {ip_str}가 삭제되었습니다.')
        return redirect('assets:ip_list')
    return render(request, 'assets/ip_confirm_delete.html', {'obj': obj})


@login_required
def ip_person_search(request):
    """사용자 검색 AJAX — IP 수정 폼에서 사용"""
    q = request.GET.get('q', '').strip()
    results = []
    if q and len(q) >= 1:
        persons = (
            Person.objects
            .select_related('department', 'company')
            .filter(
                Q(name__icontains=q) | Q(legacy_id__icontains=q)
            )
            .order_by('name')[:30]
        )
        for p in persons:
            dept = p.department.name if p.department else ''
            co = p.company.name if p.company else ''
            label = p.name
            if dept:
                label += f' ({dept})'
            if co:
                label += f' [{co}]'
            results.append({'id': p.pk, 'text': label, 'use_yn': p.use_yn})
    return JsonResponse({'results': results})


@login_required
def ip_export(request):
    qs = IpAddress.objects.select_related(
        'group', 'person', 'person__department', 'person__company'
    ).order_by('ip_int')

    q = request.GET.get('q', '').strip()
    group_id = request.GET.get('group', '')
    status = request.GET.get('status', '')
    company_id = request.GET.get('company', '')

    if q:
        qs = qs.filter(Q(ip__icontains=q) | Q(person__name__icontains=q) | Q(note__icontains=q))
    if group_id:
        qs = qs.filter(group_id=group_id)
    if status == 'used':
        qs = qs.filter(person__isnull=False)
    elif status == 'free':
        qs = qs.filter(person__isnull=True)
    if company_id:
        qs = qs.filter(person__company_id=company_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'IP 관리'

    header_fill = PatternFill(fill_type='solid', fgColor='004A98')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin', color='D8DEE9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    headers = ['IP 주소', 'IP 그룹', '상위 그룹', '사용자', '부서', '회사', '메모', '사용 시작일', '사용 종료일', '상태']
    col_widths = [18, 20, 20, 16, 20, 16, 30, 14, 14, 10]

    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    for row_idx, ip in enumerate(qs, start=2):
        row = [
            ip.ip,
            ip.group.name if ip.group else '',
            ip.group.parent_name if ip.group else '',
            ip.person.name if ip.person else '',
            ip.person.department.name if ip.person and ip.person.department else '',
            ip.person.company.name if ip.person and ip.person.company else '',
            ip.note,
            ip.start_date,
            ip.end_date,
            '사용중' if ip.person else '미사용',
        ]
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val or '')
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    log_action(request, 'export', 'IpAddress', '', f'IP 목록 Excel 내보내기 ({qs.count()}건)')

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ip_list.xlsx"'
    return response


# ──────────────────────────────────────────────
# 전화번호 관리 (stub — 작업순서 10에서 구현)
# ──────────────────────────────────────────────

@login_required
def phone_list(request):
    return render(request, 'assets/phone_list.html', {})
